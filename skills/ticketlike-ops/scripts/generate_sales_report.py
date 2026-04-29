#!/usr/bin/env python3
"""
Genera reporte Excel de ventas segmentado por tipo de ubicación.
Uso: python3 generate_sales_report.py <eventId> [output_path]
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_connect import query

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    os.system("sudo pip3 install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if len(sys.argv) < 2:
    print("Uso: python3 generate_sales_report.py <eventId> [output_path]")
    sys.exit(1)

event_id = sys.argv[1]
output = sys.argv[2] if len(sys.argv) > 2 else f"/home/ubuntu/Reporte_Evento_{event_id}.xlsx"

# Get event info
event = query("SELECT * FROM events WHERE id = %s", (event_id,))
if not event:
    print(f"Evento {event_id} no encontrado")
    sys.exit(1)
title = event[0].get('title', f'Evento {event_id}')

# Get all paid orders with seat details
orders = query("""
    SELECT DISTINCT o.id, o.confirmationCode, o.customerName, o.customerEmail,
           o.customerPhone, o.ticketType, o.quantity, o.amountCents/100 as mxn,
           o.createdAt, o.status
    FROM ticket_orders o
    WHERE o.eventId = %s AND o.status = 'paid'
    ORDER BY o.ticketType, o.createdAt
""", (event_id,))

# Get seat assignments
seats_data = query("""
    SELECT es.orderId, s.label, s.section
    FROM event_seats es
    JOIN seats s ON es.seatId = s.id
    WHERE es.eventId = %s AND es.status = 'sold'
    ORDER BY s.label
""", (event_id,))

# Map orderId → seat labels
order_seats = {}
for s in seats_data:
    oid = s['orderId']
    if oid not in order_seats:
        order_seats[oid] = []
    order_seats[oid].append(s['label'])

# Categorize
categories = {
    'BUTACAS': [],
    'MESA VIP 4': [],
    'MESA VIP 6': [],
    'SALA VIP': [],
    'OTROS': [],
}

for o in orders:
    tt = (o['ticketType'] or '').upper()
    seat_labels = order_seats.get(o['id'], [])
    o['seat_labels'] = ', '.join(seat_labels) if seat_labels else 'N/A'
    
    if 'BUTACA' in tt or 'BLEACHER' in tt:
        categories['BUTACAS'].append(o)
    elif 'MESA VIP 4' in tt or (('MESA' in tt or 'VIP' in tt) and '4' in tt):
        categories['MESA VIP 4'].append(o)
    elif 'MESA VIP 6' in tt or (('MESA' in tt or 'VIP' in tt) and '6' in tt):
        categories['MESA VIP 6'].append(o)
    elif 'SALA' in tt:
        categories['SALA VIP'].append(o)
    else:
        categories['OTROS'].append(o)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
cortesia_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ['#', 'Código', 'Cliente', 'Email', 'Teléfono', 'Asiento(s)', 'Tipo', 'Monto', 'Fecha']

for cat_name, cat_orders in categories.items():
    if not cat_orders:
        continue
    ws = wb.create_sheet(title=cat_name[:31])
    
    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for i, o in enumerate(cat_orders, 1):
        row = i + 1
        values = [
            i,
            o['confirmationCode'],
            o['customerName'],
            o['customerEmail'],
            o['customerPhone'],
            o['seat_labels'],
            o['ticketType'],
            f"${o['mxn']:,.2f}",
            str(o['createdAt'])[:16],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if o['mxn'] == 0:
                cell.fill = cortesia_fill
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, len(cat_orders) + 2))
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = min(max_len + 2, 40)

# Summary sheet
ws = wb.create_sheet(title="RESUMEN", index=0)
ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
ws.cell(row=3, column=1, value="Sección").font = Font(bold=True)
ws.cell(row=3, column=2, value="Órdenes").font = Font(bold=True)
ws.cell(row=3, column=3, value="Cortesías").font = Font(bold=True)
ws.cell(row=3, column=4, value="Con Pago").font = Font(bold=True)
ws.cell(row=3, column=5, value="Revenue").font = Font(bold=True)

row = 4
total_orders = 0
total_revenue = 0
for cat_name, cat_orders in categories.items():
    if not cat_orders:
        continue
    cortesias = sum(1 for o in cat_orders if o['mxn'] == 0)
    con_pago = len(cat_orders) - cortesias
    revenue = sum(o['mxn'] for o in cat_orders)
    ws.cell(row=row, column=1, value=cat_name)
    ws.cell(row=row, column=2, value=len(cat_orders))
    ws.cell(row=row, column=3, value=cortesias)
    ws.cell(row=row, column=4, value=con_pago)
    ws.cell(row=row, column=5, value=f"${revenue:,.2f}")
    total_orders += len(cat_orders)
    total_revenue += revenue
    row += 1

ws.cell(row=row+1, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=row+1, column=2, value=total_orders).font = Font(bold=True)
ws.cell(row=row+1, column=5, value=f"${total_revenue:,.2f}").font = Font(bold=True)

wb.save(output)
print(f"✅ Reporte generado: {output}")
print(f"   {total_orders} órdenes, ${total_revenue:,.2f} MXN")
